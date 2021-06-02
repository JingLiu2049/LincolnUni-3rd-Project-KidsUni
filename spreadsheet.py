import db
import pandas as pd
import openpyxl as op
import os
from openpyxl import load_workbook
import datetime as dt
basepath  = os.path.dirname(__file__)

def new_sheet(path, sql, name):
    book =load_workbook(path)
    writer = pd.ExcelWriter(path,engine='openpyxl')
    writer.book = book
    df = pd.read_sql(sql,db.get_conn())
    df.to_excel(writer,index=False,sheet_name=f'{name}')
    writer.save()

def excel_obj(temp_name):
    templatePath =os.path.join(basepath,'downloads',temp_name)
    bg = op.load_workbook(templatePath)
    return bg

def gen_newPath(ss_name):
    filename = f'{dt.datetime.now().year}_{ss_name}.xlsx'
    newPath = os.path.join(basepath,'downloads',filename)
    return newPath

def gen_mem_tmp(schoolid):
    # get spreadsheet template object
    bg = excel_obj('End year template.xlsx')
    sheet1 = bg['Sheet1']
    sheet2 = bg['Username and passwords']
    # generating new path 
    cur = db.getCursor()
    cur.execute("SELECT school_name FROM schools WHERE school_id = %s;",(schoolid ,))
    sch_name = cur.fetchone()[0]
    newPath = gen_newPath(f'{sch_name}_Template')

     # get coordinator information from databse and insert into spreadsheet  
    cur.execute("SELECT name, email, phone_number,username, password FROM coordinator \
        WHERE school_id  = %s;",(schoolid ,))
    coor = cur.fetchone()
    if coor:
        for i in range(0,len(coor[0:3])):
            sheet1.cell(column = 4,row = 2 + i,value = coor[0:3][i])
        coor_u_p = [coor[0],'Coordinator']+list(coor[-2:])
        for j in range(0,len(coor_u_p)):
            sheet2.cell(column = j+1, row = 3, value = coor_u_p[j])

    # get member information from databse and insert into spreadsheet
    sheet1.cell(column = 4, row =1, value = sch_name.capitalize())
    sheet2.cell(column = 1, row =1, value = sch_name.capitalize())
    sql = "SELECT member_id, first_name, last_name, gender, member_age, ethnicity, \
        continuing_new, status, passport_number, passport_date_issued, \
        ethnicity_info, teaching_research, publication_promos, social_media,gown_size, hat_size,username,password FROM \
        members WHERE school_id = %s ORDER BY member_id" % schoolid 
    cur.execute(sql)
    mem_infos = cur.fetchall()
    if mem_infos:
        for i in range(0,len(mem_infos)):
            mem_info = mem_infos[i][0:-4] 
            for j in range(0,len(mem_info)):
                sheet1.cell(column = j+1,row = 7+i,value = mem_info[j])
            u_p = list(mem_infos[i][1:3])+list(mem_infos[i][-2:])
            for k in range(0,len(u_p)):
                sheet2.cell(column = k+1, row = 7+i, value = u_p[k])
            gown_hat = mem_infos[i][-4:-2]
            for l in range(0,2):
                sheet1.cell(column = 21+l, row = 7+i, value = gown_hat[l])
    # get events and cut-off date info from databse and insert into spreadsheet
    # insert event info
    cur.execute("SELECT name, event_date, event_id FROM events WHERE EXTRACT(YEAR FROM event_date) = \
        EXTRACT(year from now());")
    events = cur.fetchall()
    if events:
        for i in range(0,len(events)):
            for j in range(0,3):
                sheet1.cell(column = 23+i, row = 4+j, value = events[i][j])
    # insert data cut-off date
    year = dt.datetime.now().year
    sheet1.cell(column = 4,row = 5,value = f'{year}-12-31')
    # change column continuing/new to continuing for members 
    for i in range(0,len(mem_infos)):
        sheet1.cell(column = 7, row=7+i, value='Continuing')
    # get previous hours info of each member from databse and insert into spreadsheet
    cur.execute("SELECT previous FROM previous WHERE school_id = %s ORDER BY member_id;",(int(schoolid),))
    previous = cur.fetchall()
    for i in range(0,len(previous)):
        sheet1.cell(column=15, row=7+i, value=previous[i][0])
    bg.save(newPath)
    #  insert new sheet
    pd_sql_event = "SELECT * FROM events WHERE EXTRACT(YEAR FROM event_date) = EXTRACT(year from now()) \
        ORDER BY event_id;"
    new_sheet(newPath,pd_sql_event,'Events')
    return newPath

def gen_mem_comp(schoolid):
        # get template object
        bg = excel_obj('End year template.xlsx')
        sheet1 = bg['Sheet1']
        sheet2 = bg['Username and passwords']

        # generating new path 
        cur = db.getCursor()
        cur.execute("SELECT school_name FROM schools WHERE school_id = %s;",(schoolid ,))
        sch_name = cur.fetchone()[0]
        newPath = gen_newPath(f'{sch_name}_Completed')

        # get events and cut-off date info from databse and insert into spreadsheet
        # insert cut-off date
        cur.execute("SELECT max(year) FROM membershours;")
        cut_date = cur.fetchone()
        sheet1.cell(column = 4,row = 5,value = cut_date[0])

        # get member information from databse and insert into spreadsheet
        sheet1.cell(column = 4, row =1, value = sch_name.capitalize())
        sheet2.cell(column = 1, row =1, value = sch_name.capitalize())
        sql = "SELECT member_id, first_name, last_name, gender, member_age, ethnicity, \
            continuing_new, status, passport_number, passport_date_issued, \
            ethnicity_info, teaching_research, publication_promos, social_media, previous,\
            gown_size, hat_size, username,password FROM members WHERE school_id = %s ORDER \
            BY member_id" % schoolid 
        cur.execute(sql)
        mem_infos = cur.fetchall()
        if mem_infos:
            for i in range(0,len(mem_infos)):
                mem_info = mem_infos[i][0:-4] 
                for j in range(0,len(mem_info)):
                    sheet1.cell(column = j+1,row = 7+i,value = mem_info[j])
                u_p = list(mem_infos[i][1:3])+list(mem_infos[i][-2:])
                for k in range(0,len(u_p)):
                    sheet2.cell(column = k+1, row = 7+i, value = u_p[k])
                gown_hat = mem_infos[i][-4:-2]
                for l in range(0,2):
                    sheet1.cell(column = 21+l, row = 7+i, value = gown_hat[l])

        # get total hours info of each member from databse and insert into spreadsheet
        cur.execute("SELECT previous FROM previous WHERE school_id = %s ORDER BY member_id;",(int(schoolid),))
        previous = cur.fetchall()
        for i in range(0,len(previous)):
            sheet1.cell(column=20, row=7+i, value=previous[i][0])
        
        # get the newest detail hours info and insert into ss
        for m in range(0,4):
            cur.execute(f"SELECT hours FROM detail_hours WHERE term = 'term{m+1}' and school_id = {schoolid} and extract(year from year) = {cut_date[0].year};")
            term_hour = cur.fetchall()
            for n in range(0,len(term_hour)):
                sheet1.cell(column=16+m,row=7+n,value= term_hour[n][0])

        # get coordinator information from databse and insert into spreadsheet  
        cur.execute("SELECT name, email, phone_number,username, password FROM coordinator \
            WHERE school_id  = %s;",(schoolid ,))
        coor = cur.fetchone()
        if coor:
            for i in range(0,len(coor[0:3])):
                sheet1.cell(column = 4,row = 2 + i,value = coor[0:3][i])
            coor_u_p = [coor[0],'Coordinator']+list(coor[-2:])
            for j in range(0,len(coor_u_p)):
                sheet2.cell(column = j+1, row = 3, value = coor_u_p[j])

        
        # insert event title
        cur.execute("SELECT name, event_date, event_id FROM events ORDER BY event_id;")
        events = cur.fetchall()
        if events:
            for i in range(0,len(events)):
                for j in range(0,3):
                    sheet1.cell(column = 23+i, row = 4+j, value = events[i][j])
        
        # get attendance info of each member and insert into spreadsheet
        for i in range(0,len(events)):
            eventid = events[i][2]
            sql = "SELECT attendance.status FROM members LEFT JOIN \
            attendance ON members.member_id = attendance.member_id WHERE members.school_id \
            = %s AND attendance.event_id = %s ORDER BY members.member_id" %(schoolid, eventid)
            cur.execute(sql)
            attends = cur.fetchall()
            for j in range(0,len(attends)):
                sheet1.cell(column = 23+i, row = 7+j, value = attends[j][0])
        pd_sql_event = "SELECT * FROM events ORDER BY event_id;"
        pd_sql_hours = "SELECT member_id, first_name, last_name, year, term, hours \
        FROM detail_hours WHERE school_id = %s" % int(schoolid)
        #  save new excel file
        bg.save(newPath)
        #  insert new sheet
        new_sheet(newPath,pd_sql_event,'Events')
        new_sheet(newPath,pd_sql_hours,'Hours')
        return newPath


def gen_dest_sheet():
    bg = excel_obj('Learning Destination template.xlsx')
    sheet1 = bg['Sheet1']

    # generating new path 
    newPath = gen_newPath('Learning Destination')

    cur = db.getCursor()
    cur.execute("SELECT * FROM destinations ORDER BY ld_id;")
    dests = cur.fetchall()
    for i in range(0,len(dests)):
        for j in range(0,len(dests[i])):
            sheet1.cell(column = j+1,row = 3+i,value = dests[i][j])

    cur.execute("SELECT DISTINCT extract(year from year) as year from paperwork ORDER BY year;")
    years = cur.fetchall()
    
    for i in range(0,len(years)):
        year = years[i][0]
        sheet1.cell(column = 21+i, row = 2, value = year)
        cur.execute("SELECT status FROM ld_paperwork where year = %s ORDER BY ld_id;",(float(year),))
        ppw_status = cur.fetchall()
        for j in range(0,len(ppw_status)):
            sheet1.cell(column = 21+i, row = 3+j, value = ppw_status[j][0])
    bg.save(newPath)
    return newPath


def gen_volun_sheet():
    bg = excel_obj('Volunteer database template.xlsx')
    sheet1 = bg['Volunteer Details']
    sheet2 = bg['Volunteer Hours']

    # generating new path 
    newPath = gen_newPath('Volunteers')

    cur = db.getCursor()
    sql ="SELECT volunteers.*, volunteerform.* FROM volunteers INNER JOIN volunteerform \
        ON volunteers.volun_id = volunteerform.volun_id ORDER BY volunteers.volun_id;"
    cur.execute(sql)
    voluns = cur.fetchall()
    cur.execute("SELECT sum(hours) AS total FROM volun_hours GROUP BY volun_id ORDER BY volun_id;")
    total = cur.fetchall()
    for i in range(0,len(voluns)):
        allData = list(voluns[i])
        allData.pop(14)
        pData = [allData[0]]+allData[6:8]
        for j in range(0,len(allData)):
            sheet1.cell(column = j+1,row = 2+i,value = allData[j])
        for k in range(0,3):
            sheet2.cell(column = k+1, row = 6+i, value=pData[k] )
        sheet2.cell(column=5,row=6+i, value=total[i][0])

    cur.execute("SELECT name, event_date, event_id FROM events ORDER BY event_id;")
    events = cur.fetchall()
    if events:
        for i in range(0,len(events)):
            for j in range(0,3):
                sheet2.cell(column = 6+i, row = 3+j, value = events[i][j])
    
    # get attendance info of each volunteer and insert into spreadsheet
        for i in range(0,len(events)):
            eventid = events[i][2]
            sql = "SELECT volun_hours.hours FROM volunteers LEFT JOIN \
            volun_hours ON volunteers.volun_id = volun_hours.volun_id WHERE volun_hours.event_id \
            = %s ORDER BY volunteers.volun_id" % eventid
            cur.execute(sql)
            attends = cur.fetchall()
            for j in range(0,len(attends)):
                sheet2.cell(column = 6+i, row = 6+j, value = attends[j][0])
    bg.save(newPath)
    return newPath


def gen_sch_temp():
    bg = excel_obj('School template.xlsx')
    sheet1 = bg['School list']

    # generating new path 
    newPath = gen_newPath('Schools_Template')

    cur = db.getCursor()
    cur.execute("SELECT * FROM sch_detail ORDER BY school_id;")
    schools = cur.fetchall()
    current_year = f'{dt.datetime.now().year}'
    cur.execute("SELECT MAX(year) FROM school_members;")
    result = cur.fetchone()
    last_year = int(result[0]) if result and int(result[0]) < int(current_year) else int(dt.datetime.now().year)-1

    cur.execute("SELECT confirm_no FROM school_members WHERE year = %s ORDER BY school_id",(last_year,))
    result = cur.fetchall()
    totals = result if result else [0]*len(schools)
    
    for i in range(0,len(schools)):
        sch_value = list(schools[i]) + [current_year] + [totals[i]]
        for j in range(0,len(sch_value)):
            sheet1.cell(column = j+1,row = 2+i,value = sch_value[j] )
    bg.save(newPath)
    return newPath

def gen_sch_comp():
    bg = excel_obj('School template.xlsx')
    sheet1 = bg['School list']

    # generating new path 
    newPath = gen_newPath('Schools_Completed')

    cur = db.getCursor()
    cur.execute("SELECT * FROM sch_detail ORDER BY school_id;")
    schools = cur.fetchall()
    cur.execute("SELECT MAX(year) FROM school_members;")
    result = cur.fetchone()
    current_year = int(result[0]) if result else int(dt.datetime.now().year)
    
    for i in range(0,len(schools)):
        sch_value = list(schools[i]) + [current_year]
        for j in range(0,len(sch_value)):
            sheet1.cell(column = j+1,row = 2+i,value = sch_value[j] )

    cur.execute("SELECT return_no, max_no, request_no, confirm_no FROM school_members WHERE year = %s ORDER BY school_id;",(current_year,))
    member_nos = cur.fetchall()
    cur.execute("SELECT confirm_no FROM school_members WHERE year = %s ORDER BY school_id",(current_year-1,))
    result = cur.fetchall()
    totals = result if result else [0]*len(member_nos)

    for i in range(0,len(member_nos)):
        mem_no_detail = [totals[i]] + list(member_nos[i])
        for j in range(0,len(mem_no_detail)):
            sheet1.cell(column = j+18,row = 2+i,value = mem_no_detail[j] )

    bg.save(newPath)
    pd_sql = "SELECT schools.school_name, school_members.*  FROM school_members INNER JOIN schools ON \
        school_members.school_id = schools.school_id ORDER BY school_id, year DESC;"
    new_sheet(newPath,pd_sql,'Previous Member Details')
    return newPath